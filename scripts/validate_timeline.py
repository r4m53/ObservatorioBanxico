from __future__ import annotations

from calendar import monthrange
from collections import Counter
from datetime import datetime
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "master" / "RAdarMonetario_Indice_Heath_FINAL_CORREGIDO_v20260729.xlsm"
DATASET = ROOT / "public" / "data" / "radar-bm.json"


def iso(value):
    return value.date().isoformat() if isinstance(value, datetime) else str(value)[:10]


def main():
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=False, keep_vba=True)
    monthly_dates = workbook.defined_names.get("MonthlyDates")
    assert monthly_dates is not None, "Falta el rango definido MonthlyDates"
    assert "Monthly_Experience!$A$3" in monthly_dates.attr_text

    expected = [
        iso(row[0])
        for row in workbook["Monthly_Experience"].iter_rows(min_row=3, values_only=True)
        if row[0]
    ]
    data = json.loads(DATASET.read_text(encoding="utf-8"))
    actual = data["timeline"]

    assert actual == expected, "La línea temporal publicada difiere de MonthlyDates"
    assert actual == sorted(actual), "Las fechas no están ordenadas"
    assert len(actual) == len(set(actual)), "Hay fechas duplicadas"
    assert all(
        int(date[8:10]) == monthrange(int(date[:4]), int(date[5:7]))[1]
        for date in actual
    ), "Hay fechas que no son fin de mes"

    boards = data["boards"]
    assignments = {
        date: [
            board
            for board in boards
            if board["start"] <= date <= board["end"]
        ]
        for date in actual
    }
    invalid = {date: matches for date, matches in assignments.items() if len(matches) != 1}
    assert not invalid, f"Fechas sin exactamente una Junta vigente: {list(invalid)[:5]}"

    expected_months = []
    year, month = map(int, actual[0][:7].split("-"))
    last_year, last_month = map(int, actual[-1][:7].split("-"))
    while (year, month) <= (last_year, last_month):
        expected_months.append(f"{year:04d}-{month:02d}")
        month += 1
        if month == 13:
            year, month = year + 1, 1
    assert [date[:7] for date in actual] == expected_months, "Existen huecos mensuales"

    evaluation_keys = Counter(
        (evaluation["person"], evaluation["date"], evaluation["origin"])
        for evaluation in data["evaluations"]
    )
    duplicates = [key for key, count in evaluation_keys.items() if count > 1]
    assert not duplicates, f"Evaluaciones duplicadas por persona/fecha/origen: {duplicates[:5]}"

    print(
        f"OK: {len(actual)} fechas, 0 meses faltantes, 0 duplicados, "
        "1 Junta vigente por fecha"
    )


if __name__ == "__main__":
    main()
