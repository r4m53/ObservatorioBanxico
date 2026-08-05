from __future__ import annotations
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import json, os, sys
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(os.environ.get("RADAR_BM_SOURCE", ROOT / "data" / "master" / "RAdarMonetario_Indice_Heath_FINAL_CORREGIDO_v20260729.xlsm"))
OUTPUT = ROOT / "public" / "data" / "radar-bm.json"
ACADEMIC_VALIDATION_OUTPUT = ROOT / "public" / "data" / "academic-validation.json"
ACADEMIC_ORDERS = {
    1: "Licenciatura",
    2: "Maestría",
    3: "Certificación",
    4: "Doctorado",
    5: "Especialidad",
    6: "Posdoctorado",
    7: "Otro",
}

def iso(v):
    return v.date().isoformat() if isinstance(v, datetime) else (str(v)[:10] if v else "")

def rows(ws, header_row=3):
    headers=[c.value for c in ws[header_row]]
    for vals in ws.iter_rows(min_row=header_row+1, values_only=True):
        if not any(v is not None for v in vals): continue
        yield {headers[i]: vals[i] for i in range(min(len(headers),len(vals))) if headers[i]}

def num(v):
    try: return round(float(v), 6)
    except (TypeError,ValueError): return None

def rebrand(value):
    """Keep generated public data aligned with the definitive project name."""
    if isinstance(value, str):
        legacy_labels = ("Observatorio" + " " + "Banxico", "Observatorio" + "Banxico", "Observatorio" + "_" + "Banxico")
        for legacy_label in legacy_labels:
            value = value.replace(legacy_label, "RAdarMonetario")
        return value
    if isinstance(value, list):
        return [rebrand(item) for item in value]
    if isinstance(value, dict):
        return {key: rebrand(item) for key, item in value.items()}
    return value
def academic_catalog(wb, people_raw):
    if "Formacion_Academica" not in wb.sheetnames:
        raise ValueError("El libro maestro no contiene la hoja Formacion_Academica")
    academic = defaultdict(list)
    issues = []
    duplicate_keys = set()
    for source_order, r in enumerate(rows(wb["Formacion_Academica"]), start=1):
        pid = str(r.get("Persona_ID") or "").strip()
        if not pid:
            issues.append({"row": source_order + 3, "type": "Persona_ID inexistente", "detail": "Registro sin Persona_ID"})
            continue
        if pid not in people_raw:
            issues.append({"row": source_order + 3, "type": "Persona_ID inexistente", "detail": pid})
            continue
        order_value = num(r.get("Orden_Grado"))
        order = int(order_value) if order_value is not None and order_value.is_integer() else None
        degree_type = str(r.get("Tipo_Grado") or "").strip()
        if order not in ACADEMIC_ORDERS:
            issues.append({"row": source_order + 3, "type": "Orden_Grado inválido", "detail": str(r.get("Orden_Grado") or "")})
            continue
        if degree_type not in ACADEMIC_ORDERS.values():
            issues.append({"row": source_order + 3, "type": "Tipo_Grado inválido", "detail": degree_type})
            continue
        if ACADEMIC_ORDERS[order] != degree_type:
            issues.append({"row": source_order + 3, "type": "Orden/Tipo inconsistente", "detail": f"{order} / {degree_type}"})
        year_value = num(r.get("Anio"))
        year = int(year_value) if year_value is not None and year_value.is_integer() else None
        fingerprint = (
            pid, order, degree_type,
            str(r.get("Programa_Normalizado") or "").strip(),
            str(r.get("Institucion_Normalizada") or "").strip(),
            year,
            str(r.get("Estado_Formacion") or "").strip(),
        )
        if fingerprint in duplicate_keys:
            issues.append({"row": source_order + 3, "type": "Registro duplicado", "detail": " | ".join(str(value or "") for value in fingerprint)})
        duplicate_keys.add(fingerprint)
        academic[pid].append({
            "order": order,
            "type": degree_type,
            "originalProgram": str(r.get("Programa_Original") or "").strip(),
            "program": str(r.get("Programa_Normalizado") or "").strip(),
            "institution": str(r.get("Institucion_Normalizada") or "").strip(),
            "country": str(r.get("Pais") or "").strip(),
            "year": year,
            "status": str(r.get("Estado_Formacion") or "").strip(),
            "sourceOrder": source_order,
            "evidenceLevel": str(r.get("Nivel_Evidencia") or "").strip(),
            "document": str(r.get("Documento") or "").strip(),
            "url": str(r.get("URL") or "").strip(),
            "consultedAt": iso(r.get("Fecha_Consulta")),
            "observations": str(r.get("Observaciones") or "").strip(),
        })
    return academic, issues

def comparador_timeline(wb):
    """Return the exact date list exposed by the Excel comparators."""
    monthly_dates = wb.defined_names.get("MonthlyDates")
    if monthly_dates is None or "Monthly_Experience!$A$3" not in monthly_dates.attr_text:
        raise ValueError("El rango MonthlyDates del Comparador no apunta a Monthly_Experience!A3")
    dates = [
        iso(values[0])
        for values in wb["Monthly_Experience"].iter_rows(min_row=3, values_only=True)
        if values[0]
    ]
    if not dates:
        raise ValueError("El rango MonthlyDates del Comparador está vacío")
    if len(dates) != len(set(dates)):
        raise ValueError("El rango MonthlyDates del Comparador contiene fechas duplicadas")
    if dates != sorted(dates):
        raise ValueError("El rango MonthlyDates del Comparador no está ordenado")
    return dates

def main():
    if not SOURCE.exists(): raise SystemExit(f"No se encontró el libro maestro: {SOURCE}")
    wb=load_workbook(SOURCE,data_only=True,read_only=False,keep_vba=True)
    criteria=[]
    for r in rows(wb["Heath_Criterios"]):
        criteria.append({"id":r["Criterio_ID"],"number":int(r["No"]),"name":r["Criterio"],"short":r["Etiqueta_Corta"],"nature":r["Naturaleza"],"definition":r["Definicion_Operativa"]})
    people_raw={r["Persona_ID"]:r for r in rows(wb["Heath_Personas"])}
    academic, academic_issues = academic_catalog(wb, people_raw)
    id_by_name={r["Nombre_Canonico"]:r["Persona_ID"] for r in people_raw.values()}
    # Known source name variants.
    aliases={"Jonathan Heath Constable":"Jonathan Heath Constable","José Gabriel Cuadra García":"José Gabriel Cuadra García"}
    career=defaultdict(list)
    for r in rows(wb["Career_History"]):
        career[r["Persona"]].append({"institution":r["Institución"],"role":r["Cargo"],"start":iso(r["Fecha inicio"]),"end":iso(r["Fecha fin"]),"description":r["Descripción breve de funciones"] or "","source":r["Fuente principal"] or "","url":r["URL"] or "","countTotal":bool(r["Cuenta_Total"]),"countMonetary":bool(r["Cuenta_Monetaria"]),"countFiscal":bool(r["Cuenta_Fiscal"])})
    profiles={r["Persona"]:r for r in rows(wb["Board_Profile"])}
    people=[]
    for pid,r in people_raw.items():
        name=r["Nombre_Canonico"]
        p=profiles.get(name,{})
        profile={str(k):num(v) if isinstance(v,(int,float)) else (iso(v) if isinstance(v,datetime) else str(v or "")) for k,v in p.items() if k!="Persona"}
        people.append({"id":pid,"name":name,"firstBoardDate":iso(r["Primera_Fecha_Junta"]),"lastBoardDate":iso(r["Ultima_Fecha_Junta"]),"active":str(r["Activo_Junta"]).lower()=="sí","biography":r["Observaciones"] or "Perfil curricular documentado en el libro maestro del RAdarMonetario.","career":career.get(name,[]),"academic":academic.get(pid,[]),"profile":profile})
    details=defaultdict(dict); detail_notes=defaultdict(dict)
    for r in rows(wb["Heath_Detalle"]):
        score=num(r["Calificacion"])
        if score is not None:
            details[r["Evaluacion_ID"]][r["Criterio_ID"]]=score
        detail_notes[r["Evaluacion_ID"]][r["Criterio_ID"]]=r["Comentario"] or ""
    evaluations=[]
    for r in rows(wb["Heath_Evaluaciones"]):
        if not r.get("Evaluacion_ID") or not r.get("Persona_ID"): continue
        evaluations.append({"id":r["Evaluacion_ID"],"personId":r["Persona_ID"],"person":r["Persona"],"date":iso(r["Fecha_Evaluacion"]),"total":num(r["Total"]),"origin":r["Origen_Evaluacion"] or "","evaluator":r["Evaluador"] or "","notes":r["Observaciones_Generales"] or "","source":r["Fuente_Principal"] or "","sourceUrl":r["URL_Fuente"] or "","scores":details.get(r["Evaluacion_ID"],{}),"scoreNotes":detail_notes.get(r["Evaluacion_ID"],{})})
    boards=[]
    for r in rows(wb["Board_Composition"]):
        members=[r.get("Gobernador"),r.get("Subgobernador 1"),r.get("Subgobernador 2"),r.get("Subgobernador 3"),r.get("Subgobernador 4")]
        members=[x for x in members if x]
        boards.append({"date":iso(r["Fecha fin"]),"start":iso(r["Fecha inicio"]),"end":iso(r["Fecha fin"]),"governor":r["Gobernador"],"deputies":members[1:],"members":members})
    exp=[]
    ws=wb["Monthly_Experience"]
    for vals in ws.iter_rows(min_row=3,values_only=True):
        if not vals[0]: continue
        exp.append({"date":iso(vals[0]),"total":num(vals[1]),"fiscal":num(vals[2]),"monetary":num(vals[3])})
    timeline=comparador_timeline(wb)
    payload=rebrand({"metadata":{"generatedAt":datetime.now().isoformat(timespec="seconds"),"source":SOURCE.name,"latestDate":timeline[-1],"timelineSource":"Excel: MonthlyDates (Comparador → Monthly_Experience!A3:A)","academicSource":"Excel: Formacion_Academica (relación exclusiva por Persona_ID)","academicValidation":{"records":sum(len(records) for records in academic.values()),"peopleWithRecords":sum(1 for records in academic.values() if records),"issues":len(academic_issues)}},"timeline":timeline,"criteria":criteria,"people":people,"boards":boards,"evaluations":evaluations,"experience":exp})
    OUTPUT.parent.mkdir(parents=True,exist_ok=True)
    OUTPUT.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8")
    ACADEMIC_VALIDATION_OUTPUT.write_text(json.dumps({"source":SOURCE.name,"generatedAt":datetime.now().isoformat(timespec="seconds"),"issues":academic_issues},ensure_ascii=False,indent=2),encoding="utf-8")
    print(f"OK {OUTPUT} | {len(timeline)} fechas | {len(people)} personas | {len(boards)} periodos de junta | {len(evaluations)} evaluaciones | {len(exp)} meses | {sum(len(records) for records in academic.values())} formaciones | {len(academic_issues)} incidencias académicas")

if __name__=="__main__": main()
